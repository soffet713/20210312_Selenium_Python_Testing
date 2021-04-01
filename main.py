from selenium import webdriver
from selenium.webdriver.common.by import By
from datetime import date, datetime, timedelta
import openpyxl as opxl
from selenium.webdriver.common.keys import Keys
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import time


def str_to_int_or_float(value):
    if isinstance(value, bool):
        return value
    try:
        return int(value)
    except ValueError:
        try:
            return float(value)
        except ValueError:
            return value


driver = webdriver.Firefox()
driver.get("https://www.flalottery.com/site/megaMillions")
driver.maximize_window()
time.sleep(3)

fulldate = driver.find_element(By.XPATH, '//div[@class="gamePageNumbers"]/child::p[2]').text
convtdate = datetime.strptime(fulldate, "%A, %B %d, %Y")

day = convtdate.weekday()

if day == 1:
    diff = 4
elif day == 4:
    diff = 3

currdate = convtdate.strftime("%m%d%Y")
today = date.today().strftime("%Y%m%d")
wb = opxl.Workbook()
fn = today + '_Lotnums.xlsx'
wb.save(fn)

thin_border = Border(left=Side(style='thin'),
                  right=Side(style='thin'),
                  bottom=Side(style='thin'),
                  top=Side(style='thin'))

ws = wb.active
ws.title = "FLMM"
c = ws['A1']
c.font = Font(size=24,
              color='FF0000',
              bold=True,
              underline='single',
              italic=True,
              outline=True)
c.fill = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
c.alignment = Alignment(horizontal='center')
c.border = thin_border
ws['A1'].value = 'Florida Mega Millions'
ws.merge_cells('A1:I1')
col = 0
second_row = ['A2', 'B2', 'C2', 'D2', 'E2', 'F2', 'G2', 'H2', 'I2']
sr_values = ['Date', 'First Number', 'Second Number', 'Third Number', 'Fourth Number',
             'Fifth Number', 'Mega Ball Number', 'Megaplier', 'Prize']
for cell in second_row:
    ws[cell].font = Font(underline='single')
    ws[cell].value = sr_values[col]
    ws[cell].border = thin_border
    ws.column_dimensions[cell[0:1]].width = 19
    col = col + 1
wb.save(fn)


def getmmnumbs(newdate, dif, row):
    balls = driver.find_elements(By.XPATH, '//div[@class="gamePageBalls"]/child::p[1]'
                                           '/child::span[contains(@class, "balls")]')
    datestring = newdate[0:2] + '/' + newdate[2:4] + '/' + newdate[4:]
    if newdate != currdate:
        driver.find_element_by_id('singleDate').send_keys(datestring)
        driver.find_element_by_id('singleDate').send_keys(Keys.ENTER)
        driver.find_element_by_id('submit').click()
        time.sleep(3)
        balls = driver.find_elements(By.XPATH, '//div[@class="winningNumbers"]'
                                               '/child::span[contains(@class, "balls")]')
    cells = ['A' + str(row), 'B' + str(row), 'C' + str(row), 'D' + str(row), 'E' + str(row),
             'F' + str(row), 'G' + str(row), 'H' + str(row), 'I' + str(row)]
    ws[cells[0]].value = datestring
    ws[cells[0]].border = thin_border
    x = 1
    for ball in balls:
        if x <= 7:
            ws[cells[x]].value = ball.text
            ws[cells[x]].border = thin_border
        x = x + 1

    prize_amount = driver.find_element(By.XPATH, '//table[@class="style1 games"]/child::tbody[1]/child::tr'
                                                 '/child::td[@class="column3"]')
    ws[cells[8]].value = prize_amount.text
    ws[cells[8]].border = thin_border
    nextdate = datetime.strptime(newdate, "%m%d%Y") - timedelta(days=dif)
    if dif == 3:
        dif = 4
    elif dif == 4:
        dif = 3
    if nextdate >= datetime.strptime('05032016', "%m%d%Y"):
        row = row + 1
        cellrange = "A2:I" + str(row)
        ws.auto_filter.ref = cellrange
        getmmnumbs(nextdate.strftime("%m%d%Y"), dif, row)


getmmnumbs(currdate, diff, 3)

driver.close()
wb.save(fn)

driver1 = webdriver.Firefox()
driver1.get("https://www.flalottery.com/site/powerball")
driver1.maximize_window()
time.sleep(3)

fulldate1 = driver1.find_element(By.XPATH, '//div[@class="gamePageNumbers"]/child::p[2]').text
convtdate1 = datetime.strptime(fulldate1, "%A, %B %d, %Y")

day = convtdate1.weekday()

if day == 2:
    diff = 4
elif day == 5:
    diff = 3

currdate = convtdate1.strftime("%m%d%Y")

wb.create_sheet('PWRBLL')
wb.active = wb['PWRBLL']
ws1 = wb.active
c = ws1['A1']
c.font = Font(size=24,
              color='000000',
              bold=True,
              underline='single',
              italic=True,
              outline=True)
c.fill = PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')
c.alignment = Alignment(horizontal='center')
c.border = thin_border
ws1['A1'].value = 'Power Ball'
ws1.merge_cells('A1:I1')
col = 0
second_row = ['A2', 'B2', 'C2', 'D2', 'E2', 'F2', 'G2', 'H2', 'I2']
sr_values = ['Date', 'First Number', 'Second Number', 'Third Number', 'Fourth Number',
             'Fifth Number', 'PowerBall Number', 'Power Play', 'Prize']
for cell in second_row:
    ws1[cell].font = Font(underline='single')
    ws1[cell].value = sr_values[col]
    ws1[cell].border = thin_border
    ws1.column_dimensions[cell[0:1]].width = 19
    col = col + 1
wb.save(fn)


def getpbnumbs(newdate, dif, row):
    balls = driver1.find_elements(By.XPATH, '//div[@class="gamePageBalls"]/child::p[1]'
                                            '/child::span[contains(@class, "balls")]')
    datestring = newdate[0:2] + '/' + newdate[2:4] + '/' + newdate[4:]
    if newdate != currdate:
        driver1.find_element_by_id('singleDate').send_keys(datestring)
        driver1.find_element_by_id('singleDate').send_keys(Keys.ENTER)
        driver1.find_element_by_id('submit').click()
        time.sleep(3)
        balls = driver1.find_elements(By.XPATH, '//div[@class="winningNumbers"]'
                                                '/child::span[contains(@class, "balls")]')
    cells = ['A' + str(row), 'B' + str(row), 'C' + str(row), 'D' + str(row), 'E' + str(row),
             'F' + str(row), 'G' + str(row), 'H' + str(row), 'I' + str(row)]
    ws1[cells[0]].value = datestring
    ws1[cells[0]].border = thin_border
    x = 1
    for ball in balls:
        if x <= 7:
            ws1[cells[x]].value = ball.text
            ws1[cells[x]].border = thin_border
        x = x + 1

    prize_amount = driver1.find_element(By.XPATH, '//table[@class="style1 games"]/child::tbody[1]/child::tr'
                                                  '/child::td[@class="column3"]')
    ws1[cells[8]].value = prize_amount.text
    ws1[cells[8]].border = thin_border
    nextdate = datetime.strptime(newdate, "%m%d%Y") - timedelta(days=dif)
    if dif == 3:
        dif = 4
    elif dif == 4:
        dif = 3
    if nextdate >= datetime.strptime('05042016', "%m%d%Y"):
        row = row + 1
        cellrange = "A2:I" + str(row)
        ws1.auto_filter.ref = cellrange
        getpbnumbs(nextdate.strftime("%m%d%Y"), dif, row)


getpbnumbs(currdate, diff, 3)

driver1.close()
wb.save(fn)
